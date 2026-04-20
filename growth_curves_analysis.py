import datetime
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from openpyxl import load_workbook
from pathlib import Path


# ----------------------------------------------------------------------
# 1. LOAD & PARSE
# ----------------------------------------------------------------------

filepath = input('Path to Excel file: ').strip().strip('"').strip("'")
wb = load_workbook(filepath, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

# --- Layout: build {well: sample_name} ---
ROW_LETTERS = list('ABCDEFGH')
well_names = {}
col_numbers = []
current_letter = None

for row in rows:
    if row[0] == 'Layout':
        continue
    if row[1] == 'Time':
        break
    if row[2] == 1:
        col_numbers = list(row[2:14])
        continue
    if row[1] in ROW_LETTERS:
        current_letter = row[1]
        continue
    if current_letter and col_numbers:
        for i, col in enumerate(col_numbers):
            name = row[2 + i]
            if name is not None:
                well_names[f'{current_letter}{col}'] = name
        current_letter = None

# --- Data: build DataFrame (time_h + one column per active well) ---
header_idx = next(i for i, row in enumerate(rows) if row[1] == 'Time')
header = rows[header_idx]

well_cols = {header[j]: j for j in range(3, len(header))
             if header[j] in well_names}

records = []
for row in rows[header_idx + 1:]:
    t = row[1]
    if not isinstance(t, datetime.time):
        break
    if t == datetime.time(0, 0):
        continue
    hours = t.hour + t.minute / 60 + t.second / 3600
    record = {'time_h': hours}
    for well, col_idx in well_cols.items():
        record[well] = row[col_idx]
    records.append(record)

df = pd.DataFrame(records)


# ----------------------------------------------------------------------
# 2. SUMMARY
# ----------------------------------------------------------------------

conditions = {}
for well, name in well_names.items():
    if well in df.columns:
        conditions.setdefault(name, []).append(well)

print('\nAvailable conditions:')
condition_list = list(conditions.items())
for idx, (name, wells) in enumerate(condition_list):
    print(f'  [{idx}] {name:<20} — wells: {", ".join(wells)}')


# ----------------------------------------------------------------------
# 3. FILTER
# ----------------------------------------------------------------------

# --- Condition-level filter ---
cond_input = input('\nConditions to plot (e.g. 0,1,4 — or Enter for all): ').strip()
if cond_input:
    selected_indices = [int(x) for x in cond_input.split(',')]
    selected_conditions = {name: wells for i, (name, wells)
                           in enumerate(condition_list)
                           if i in selected_indices}
else:
    selected_conditions = dict(condition_list)

# --- Well-level filter ---
all_selected_wells = [w for wells in selected_conditions.values() for w in wells]
print(f'\nSelected wells: {", ".join(all_selected_wells)}')

well_input = input('Wells to exclude (e.g. A1,A3 — or Enter to keep all): ').strip()
if well_input:
    exclude = [w.strip() for w in well_input.split(',')]
    selected_conditions = {
        name: [w for w in wells if w not in exclude]
        for name, wells in selected_conditions.items()
    }
    selected_conditions = {name: wells for name, wells
                           in selected_conditions.items() if wells}

# Filter DataFrame columns
wells_to_plot = [w for wells in selected_conditions.values() for w in wells]
df_filtered = df[['time_h'] + wells_to_plot]


# ----------------------------------------------------------------------
# 4. PLOT
# ----------------------------------------------------------------------

title = input('\nPlot title (Enter to use filename): ').strip()

n_conditions = len(selected_conditions)
fig, axes = plt.subplots(1, n_conditions, figsize=(6 * n_conditions, 5))

# If only one condition, axes is not a list — wrap it so the loop works
if n_conditions == 1:
    axes = [axes]

for ax, (condition, wells) in zip(axes, selected_conditions.items()):
    colors = cm.tab20(np.linspace(0, 1, len(wells)))

    for color, well in zip(colors, wells):
        ax.plot(
            df_filtered['time_h'], df_filtered[well],
            color=color, linewidth=1.2, alpha=0.75,
            label=well,
        )

    ax.set_title(condition, fontsize=12)
    ax.set_xlabel('Time (hours)', fontsize=11)
    ax.set_ylabel('OD$_{600}$', fontsize=11)
    ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left', fontsize=8)
    ax.grid(True, linestyle='--', alpha=0.4)

fig.suptitle(title or Path(filepath).stem, fontsize=13)
plt.tight_layout(rect=[0, 0, 1, 0.95])

outfile = input('Save as (e.g. plot.png) or Enter to display: ').strip()
if outfile:
    fig.savefig(outfile, dpi=150, bbox_inches='tight')
    print(f'Saved to {outfile}')
else:
    plt.show()
